"""
API Views for Student functionality in admin panel
"""
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from .services import StudentService
from .serializers import StudentSerializer, ClassroomSerializer, BulkStudentUploadSerializer
import math # Import math để tính total_pages
import openpyxl
from datetime import datetime
import re
from io import BytesIO
from django.http import HttpResponse

@api_view(['GET'])
def get_all_students(request):
    """
    Get a paginated list of all students.
    Accepts 'page' and 'limit' query parameters.
    """
    # 1. Lấy tham số page và limit từ query params
    # Mặc định page=1, limit=10 nếu không được cung cấp
    try:
        page = int(request.query_params.get('page', 1))
        limit = int(request.query_params.get('limit', 10))
        if page < 1: page = 1
        if limit < 1: limit = 10
        # Có thể thêm giới hạn max cho limit nếu muốn (ví dụ: max=100)
        # if limit > 100: limit = 100 
    except ValueError:
        return Response({'error': 'Page and limit must be integers'}, status=status.HTTP_400_BAD_REQUEST)

    # 2. Gọi service với tham số phân trang
    result = StudentService.get_all_students(page=page, limit=limit)
    
    if result['success']:
        # 3. Serialize dữ liệu trả về từ service
        serializer = StudentSerializer(result['data'], many=True)
        
        # 4. Tính toán thông tin phân trang
        total_count = result['total_count']
        total_pages = math.ceil(total_count / limit)
        
        # 5. Tạo cấu trúc response chuẩn cho phân trang
        paginated_response = {
            'count': total_count,
            'total_pages': total_pages,
            'current_page': page,
            'limit': limit,
            'results': serializer.data # Dữ liệu thực tế nằm trong 'results'
        }
        return Response(paginated_response, status=status.HTTP_200_OK)
        
    return Response({'error': result['error']}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def get_student_by_id(request, student_id):
    """Get student details by ID"""
    result = StudentService.get_student_by_id(student_id)
    
    if result['success']:
        serializer = StudentSerializer(result['data'])
        return Response(serializer.data, status=status.HTTP_200_OK)
    return Response({'error': result['error']}, status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
def create_student(request):
    """Create a new student"""
    serializer = StudentSerializer(data=request.data)
    if serializer.is_valid():
        result = StudentService.create_student(serializer.validated_data)
        if result['success']:
            # Serialize the returned data
            response_serializer = StudentSerializer(result['data'][0])  # Get first item since insert returns array
            return Response(response_serializer.data, status=status.HTTP_201_CREATED)
        return Response({'error': result['error']}, status=status.HTTP_400_BAD_REQUEST)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT'])
def update_student(request, student_id):
    """Update student information"""
    # First check if student exists
    get_result = StudentService.get_student_by_id(student_id)
    if not get_result['success']:
        return Response({'error': 'Student not found'}, status=status.HTTP_404_NOT_FOUND)
    
    serializer = StudentSerializer(data=request.data, partial=True)
    if serializer.is_valid():
        result = StudentService.update_student(student_id, serializer.validated_data)
        if result['success']:
            response_serializer = StudentSerializer(result['data'][0])  # Get first item since update returns array
            return Response(response_serializer.data, status=status.HTTP_200_OK)
        return Response({'error': result['error']}, status=status.HTTP_400_BAD_REQUEST)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['DELETE'])
def delete_student(request, student_id):
    """Soft delete a student"""
    result = StudentService.delete_student(student_id)
    
    if result['success']:
        return Response(status=status.HTTP_204_NO_CONTENT)
    return Response({'error': result['error']}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def filter_students(request):
    """Filter students based on query parameters"""
    filters = {}
    if 'level_id' in request.query_params:
        filters['level_id'] = request.query_params['level_id']
    if 'search' in request.query_params:
        filters['search'] = request.query_params['search']
    
    result = StudentService.filter_students(filters)
    
    if result['success']:
        serializer = StudentSerializer(result['data'], many=True)
        return Response(serializer.data, status=status.HTTP_200_OK) # Trả về list như cũ
    return Response({'error': result['error']}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def get_active_classrooms_view(request):
    """
    API endpoint để lấy danh sách các lớp học đang hoạt động.
    """
    result = StudentService.get_active_classrooms()

    if result['success']:
        # Sử dụng ClassroomSerializer
        serializer = ClassroomSerializer(result['data'], many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    return Response({'error': result.get('error', 'Unknown error')}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
def bulk_upload_students(request):
    """
    API endpoint để upload file Excel và tạo hàng loạt học viên.
    Expects multipart/form-data với file field 'excel_file'.
    """
    if 'excel_file' not in request.FILES:
        return Response({'error': 'Không tìm thấy file Excel'}, status=status.HTTP_400_BAD_REQUEST)
    
    excel_file = request.FILES['excel_file']
    
    # Kiểm tra định dạng file
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        return Response({'error': 'File phải có định dạng .xlsx hoặc .xls'}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        # Đọc file Excel
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        
        # Đọc header (dòng đầu tiên)
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value.lower().strip() if cell.value else '')
        
        # Mapping các tên cột có thể có trong Excel (chỉ 3 trường bắt buộc)
        column_mapping = {
            'họ tên học viên': 'full_name',
            'họ tên': 'full_name',
            'tên': 'first_name',
            'họ': 'last_name',
            'mật khẩu': 'password',
            'password': 'password',
            'lớp học': 'classroom_code',
            'mã lớp': 'classroom_code',
            'classroom_code': 'classroom_code',
        }
        
        # Tìm index của các cột cần thiết
        column_indices = {}
        for idx, header in enumerate(headers):
            if header in column_mapping:
                column_indices[column_mapping[header]] = idx
        
        # Kiểm tra các trường bắt buộc (chỉ 3 trường: Mã lớp, Họ tên học viên, Mật khẩu)
        missing_fields = []
        
        # Kiểm tra nếu có full_name thì không cần first_name và last_name riêng
        has_full_name = 'full_name' in column_indices
        if not has_full_name:
            if 'first_name' not in column_indices:
                missing_fields.append('Tên (hoặc Họ tên học viên)')
            if 'last_name' not in column_indices:
                missing_fields.append('Họ (hoặc Họ tên học viên)')
        
        if 'password' not in column_indices:
            missing_fields.append('Mật khẩu')
        if 'classroom_code' not in column_indices:
            missing_fields.append('Mã lớp')
        
        if missing_fields:
            return Response({
                'error': f'Thiếu các cột bắt buộc: {", ".join(missing_fields)}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Đọc dữ liệu từ các dòng tiếp theo
        students_data = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            # Bỏ qua dòng trống
            if all(cell.value is None or str(cell.value).strip() == '' for cell in row):
                continue
            
            student_data = {}
            
            # Xử lý full_name nếu có (tách thành first_name và last_name)
            if has_full_name:
                full_name = row[column_indices['full_name']].value
                if full_name:
                    name_parts = str(full_name).strip().split()
                    if len(name_parts) >= 2:
                        student_data['last_name'] = name_parts[0]
                        student_data['first_name'] = ' '.join(name_parts[1:])
                    elif len(name_parts) == 1:
                        student_data['last_name'] = name_parts[0]
                        student_data['first_name'] = ''
            else:
                # Lấy first_name và last_name riêng
                if 'first_name' in column_indices:
                    student_data['first_name'] = str(row[column_indices['first_name']].value or '').strip()
                if 'last_name' in column_indices:
                    student_data['last_name'] = str(row[column_indices['last_name']].value or '').strip()
            
            # Chỉ lấy 3 trường bắt buộc: password và classroom_code
            if 'password' in column_indices:
                student_data['password'] = str(row[column_indices['password']].value or '').strip()
            
            if 'classroom_code' in column_indices:
                student_data['classroom_code'] = str(row[column_indices['classroom_code']].value or '').strip()
            
            # Validate dữ liệu cơ bản
            if not student_data.get('first_name') or not student_data.get('last_name'):
                continue  # Bỏ qua dòng không có tên
            
            if not student_data.get('password'):
                continue  # Bỏ qua dòng không có mật khẩu
            
            if not student_data.get('classroom_code'):
                continue  # Bỏ qua dòng không có mã lớp
            
            students_data.append(student_data)
        
        if not students_data:
            return Response({'error': 'Không có dữ liệu hợp lệ trong file Excel'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate dữ liệu bằng serializer
        serializer = BulkStudentUploadSerializer(data=students_data, many=True)
        if not serializer.is_valid():
            return Response({
                'error': 'Dữ liệu không hợp lệ',
                'details': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Gọi service để tạo hàng loạt học viên
        result = StudentService.bulk_create_students_from_excel(serializer.validated_data)
        
        if result['success_count'] > 0:
            return Response({
                'success': True,
                'message': f'Đã tạo thành công {result["success_count"]}/{result["total"]} học viên',
                'data': result['data'],
                'errors': result['errors'],
                'total': result['total'],
                'success_count': result['success_count'],
                'error_count': result['error_count']
            }, status=status.HTTP_201_CREATED)
        else:
            return Response({
                'success': False,
                'error': 'Không thể tạo học viên nào',
                'errors': result['errors'],
                'total': result['total'],
                'success_count': result['success_count'],
                'error_count': result['error_count']
            }, status=status.HTTP_400_BAD_REQUEST)
        
    except Exception as e:
        print(f"Lỗi khi xử lý file Excel: {e}")
        return Response({
            'error': f'Lỗi khi xử lý file Excel: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def bulk_upload_template(request):
    """
    Tạo và trả về file Excel mẫu để admin tải xuống và điền dữ liệu đúng cột.
    """
    # Tạo workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'

    # Header chỉ có 3 trường bắt buộc
    headers = [
        'Mã lớp',
        'Họ tên học viên',
        'Mật khẩu',
    ]
    ws.append(headers)

    # Hàng ví dụ (optional)
    example_rows = [
        ['2025N312', 'Đoàn Huyền Trân', '123456'],
        ['2025N312', 'Trần Thành Nhân', '123456'],
        ['2025N312', 'Trần Tùng Dương', '123456'],
    ]
    for row in example_rows:
        ws.append(row)

    # Autosize đơn giản
    for col in ws.columns:
        max_length = 12
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max_length + 2

    # Ghi ra buffer
    tmp = BytesIO()
    wb.save(tmp)
    tmp.seek(0)

    # Trả response tải xuống
    response = HttpResponse(
        tmp.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="student_upload_template.xlsx"'
    return response